import openpyxl as op

#workbook 객체 생성
filepath = r"C:\Users\owc1985\Desktop\code\autoTxt\d_list.xlsx"
wb = op.load_workbook(filepath)

#worksheet 객체 생성
ws = wb.active


# txt파일 준비
file_url = r'C:\Users\owc1985\Desktop\code\autoTxt\test.txt'
f=open(r'C:\Users\owc1985\Desktop\code\autoTxt\test.txt')

lines=f.readlines()

string = lines
for i in range(len(string)):
  ws.cell(row=i+1, column=1).value = string[i].strip()

wb.save("d_list.xlsx")

# #row, column 숫자로 접근하기
# ws.cell(row=2, column=6).value = lines

# #파일 저장 : 저장해야 결과가 반영된다. 파일명은 result로 저장
# wb.save("result.xlsx")
